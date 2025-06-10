import openpyxl.cell
import pandas as pd
from dataclasses import dataclass
from math import isnan
import random as r
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import datetime as dt


@dataclass
class ExcelProcessor:

    def _select_workbook(self, file_name: str, sheet_name: str = "FORM-16") -> None:
        """
        Opens an existing Excel workbook for Form-16 using openpyxl and selects the specified worksheet.

        This method loads the Excel file provided by `file_name` and sets the `self.workbook` attribute
        to the loaded workbook object. It then selects the worksheet specified by `sheet_name` (defaulting
        to "FORM-16") and assigns it to the `self.ws` attribute for further processing.

        Args:
            file_name (str): The path to the Excel file to open. This should be a valid .xlsx file.
            sheet_name (str, optional): The name of the worksheet to select from the workbook.
            Defaults to "FORM-16".

        Raises:
            FileNotFoundError: If the specified Excel file does not exist.
            KeyError: If the specified worksheet name does not exist in the workbook.

        Side Effects:
            Sets self.workbook to the loaded openpyxl workbook.
            Sets self.ws to the selected worksheet within the workbook.
        """
        self.workbook = openpyxl.load_workbook(file_name)
        self._select_worksheet(self.workbook, sheet_name)

    def _select_worksheet(self, workbook: openpyxl.Workbook, sheet_name: str) -> None:
        """
        Selects a worksheet from an openpyxl workbook by name and assigns it to self.ws.

        This method takes an openpyxl Workbook object and a worksheet name, and sets the
        self.ws attribute to the corresponding worksheet. It is used internally after loading
        a workbook to prepare for further processing or data manipulation.

        Args:
            workbook (openpyxl.Workbook): The loaded Excel workbook object.
            sheet_name (str): The name of the worksheet to select.

        Raises:
            KeyError: If the specified worksheet name does not exist in the workbook.

        Side Effects:
            Sets self.ws to the selected worksheet within the workbook.
        """
        try:
            self.ws = self.workbook[sheet_name]
            # Sheet exists, use it
        except KeyError:
            # Sheet doesn't exist
            self.ws = self.workbook.create_sheet(sheet_name, 0)

    def _apply_style(self, cell, style_key: str) -> None:
        """
        Applies a predefined style from self.s to a given cell.

        Args:
            cell (openpyxl.cell.cell.Cell): The cell to apply the style to.
            style_key (str): The key of the style in self.s to apply.

        Raises:
            KeyError: If the style_key does not exist in self.s.
        """
        if style_key not in self.s:
            raise KeyError(f"Style '{style_key}' not found in self.s")

        style = self.s[style_key]
        for attribute, value in style.items():
            setattr(cell, attribute, value)

    def _add_formats(self) -> None:
        """
        Defines commonly used cell styles for formatting Excel sheets and stores them in self.s.

        Side Effects:
            Sets self.s to a dictionary containing predefined cell styles.
        """
        # Define font styles

        bold_font = Font(name="calibri", bold=True, size=16)
        blue_font = Font(name="calibri", size=18, bold=True, color="FFFFFF")
        normal_font = Font(name="calibri", size=12)
        big_font = Font(name="calibri", bold=True, size=26)
        black_font = Font(name="calibri", color="FFFFFF", bold=True, size=16)
        black_font_h = Font(name="calibri", color="FFFFFF", bold=True, size=26)

        # Define alignment styles
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )

        # Define border styles
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Define fill styles for background colors
        dark_red_fill = PatternFill(
            start_color="FF0066", end_color="FF0066", fill_type="solid"
        )
        medium_red_fill = PatternFill(
            start_color="FF3399", end_color="FF3399", fill_type="solid"
        )
        light_red_fill = PatternFill(
            start_color="FF6699", end_color="FF6699", fill_type="solid"
        )
        black_fill = PatternFill(
            start_color="262626", end_color="262626", fill_type="solid"
        )
        blue_fill = PatternFill(
            start_color="002060", end_color="002060", fill_type="solid"
        )
        green_fill = PatternFill(
            start_color="00B050", end_color="00B050", fill_type="solid"
        )

        # Store styles in self.s
        self.s = {
            "blank": {"font": normal_font, "alignment": center_alignment},
            "blank_bold": {"font": bold_font, "alignment": center_alignment},
            "dark_red": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": dark_red_fill,
                "border": thin_border,
            },
            "medium_red": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": medium_red_fill,
                "border": thin_border,
            },
            "light_red": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": light_red_fill,
                "border": thin_border,
            },
            "black": {
                "font": black_font,
                "alignment": center_alignment,
                "fill": black_fill,
                "border": thin_border,
            },
            "blue": {
                "font": blue_font,
                "alignment": center_alignment,
                "fill": blue_fill,
                "border": thin_border,
            },
            "green": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": green_fill,
                "border": thin_border,
            },
            "green_h": {
                "font": big_font,
                "alignment": center_alignment,
                "fill": green_fill,
                "border": thin_border,
            },
            "dark_red_h": {
                "font": big_font,
                "alignment": center_alignment,
                "fill": dark_red_fill,
                "border": thin_border,
            },
            "medium_red_h": {
                "font": bold_font,
                "alignment": center_alignment,
                "fill": medium_red_fill,
                "border": thin_border,
            },
            "light_red_h": {
                "font": big_font,
                "alignment": center_alignment,
                "fill": light_red_fill,
                "border": thin_border,
            },
            "black_h": {
                "font": black_font_h,
                "alignment": center_alignment,
                "fill": black_fill,
                "border": thin_border,
            },
        }

    def _set_worksheet_dimensions(self):
        """
        Set all rows to height 43 and all columns to width 26.1
        """
        from openpyxl.utils import get_column_letter

        # Set row heights for all existing rows plus extra buffer
        max_row = max(self.ws.max_row, 100)  # At least 100 rows
        for row in range(1, max_row + 1):
            self.ws.row_dimensions[row].height = 43

        # Set column widths for all existing columns plus extra buffer
        max_col = max(self.ws.max_column, 26)  # At least 26 columns (A-Z)
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            self.ws.column_dimensions[col_letter].width = (
                26.1 + 0.71
            )  # type: ignore # Some Randome shit decreases the width by 0.71

        # Set defaults for any new rows/columns that might be added later
        self.ws.sheet_format.defaultRowHeight = 43
        self.ws.sheet_format.defaultColWidth = (
            26.1 + 0.71
        )  # Some Randome shit decreases the width by 0.71
        self.ws.sheet_format.customHeight = False  # Don't auto-adjust heights

    def _set_default_style(self, max_rows: int = 200, max_cols: int = 50) -> None:
        """
        Apply default 'blank' style to all cells in the worksheet
        Args:
            max_rows: Number of rows to style (default 200)
            max_cols: Number of columns to style (default 50)
        """
        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                cell = self.ws.cell(row=row, column=col)
                self._apply_style(cell, "blank")

    def set(
        self,
        cell: openpyxl.cell.Cell,
        value: str | int | float | dt.datetime,
        style: str,
        type: str = "general",
    ) -> None:
        """
        Sets the cell with value, style, and optionally cell type.

        Args:
            cell (openpyxl.cell.Cell): The cell to set.
            value (str | int | float): The value to assign.
            style (str): The style key to apply.
            type (str, optional): The type of the cell ('general', 'date', 'text', etc.). Defaults to 'general'.
        """
        cell.value = value
        self._apply_style(cell, style)

        # Set cell number format based on type
        if type == "date":
            cell.number_format = "dd/mm/yyyy"
        elif type == "text":
            cell.number_format = "@"
        elif type == "general":
            cell.number_format = "General"
        # You can add more types/formats as needed

    def make_dashboard(self, file_name: str, df: pd.DataFrame, sheet_name: str = "crypto") -> None:
        """
        Creates the Dashboard for crypto Calculations
        """

        self._select_workbook(file_name, sheet_name)

        self._add_formats()

        self._set_worksheet_dimensions()
        self._set_default_style()

        self.ws.merge_cells("A1:J1")
        crypto_details = self.ws["A1"]
        self.set(crypto_details, "Crypto Details", "black_h")

        self.ws.merge_cells("A2:B2")
        source = self.ws["A2"]
        self.set(source, "Source", "dark_red")

        doa = self.ws["C2"]
        self.set(doa, "Date of Acquisition", "medium_red")

        dot = self.ws["D2"]
        self.set(dot, "Date of Transfer", "light_red")

        coa = self.ws["E2"]
        self.set(coa, "Cost of Acquisition", "dark_red")

        cr = self.ws["F2"]
        self.set(cr, "Consideration Recived", "medium_red")

        cg = self.ws["g2"]
        self.set(cg, "Capital Gain", "blue")

        self.ws.merge_cells("H2:J2")
        msg = self.ws["H2"]
        self.set(msg, "Accha Khasa LOSS Hua Hai Bhai Saab", "black")

        """################# Total Capital Gain #################"""

        self.ws.merge_cells("H3:J3")
        tcg_title = self.ws["H3"]
        self.set(tcg_title, "Total Capital Gain", "blue")

        self.ws.merge_cells("H4:J7")
        tcg_value = self.ws["H4"]

        """################# Total Cost of Acquisition #################"""

        self.ws.merge_cells("H8:J8")
        tcoa_title = self.ws["H8"]
        self.set(tcoa_title, "Total Cost of Acquisition", "blue")

        self.ws.merge_cells("H9:J10")
        tcoa_value = self.ws["H9"]

        """################# Total Consideration Recived #################"""

        self.ws.merge_cells("H11:J11")
        tcr_title = self.ws["H11"]
        self.set(tcr_title, "Total Consideration Recived", "blue")

        self.ws.merge_cells("H12:J14")
        tcr_value = self.ws["H12"]

        """################# Evaluationg Crypto Data #################"""

        # Rename columns for consistency
        df = df.rename(
            columns={
            "Information Source": "source",
            "Date of Payment/Credit": "date_of_transfer",
            "Amount Paid/Credited - Reported by Source": "consideration_received",
            }
        )

        """################# Inserting Crypto Data #################"""

        # Inserting Details

        df_values = df.values
        for i, data in enumerate(df_values):

            print(f"({i}) \t {data[0]} \t {data[1].date()} \t {data[2]}")
            # Inserting Source Detail
            src_cell_index = "A"+ str(3+i)
            self.ws.merge_cells(src_cell_index+ ":"+ "B"+ str(3+i))
            src_cell = self.ws[src_cell_index]
            self.set(src_cell,data[0],"blank")

            # Inserting Date of Transfer
            dot_cell_index = "D" + str(3+i)
            dot_cell = self.ws[dot_cell_index]
            self.set(dot_cell,data[1],"blank","date")

            # Inserting Consideration Received
            cr_cell_index = "F" + str(3+i)
            cr_cell = self.ws[cr_cell_index]
            cr_value = data[2]
            self.set(cr_cell,cr_value,"blank")

            """################# Generating Crypto Details #################"""

            doa_cell = "C"+ str(3+i)

            # Generating Date of Acquisition
            # Generate a random date between 01-04-2024 and data[1]
            start_date = dt.datetime(2024, 4, 1)
            end_date = data[1]
            if end_date <= start_date:
                random_date = start_date
            else:
                delta = end_date - start_date
                random_days = r.randint(0, delta.days)
                random_date = start_date + dt.timedelta(days=random_days)

            doa_cell_obj = self.ws[doa_cell]
            self.set(doa_cell_obj, random_date, "blank", "date")

            # Generate Cost of Acquisition (coa) - mostly greater than consideration_received (data[2])
            # 80% chance to be greater, 20% chance to be less or equal
            if r.random() < 0.8:
                # Greater: add 5% to 30% random premium
                increment = r.uniform(0.05, 0.3)
                coa_value = round(data[2] * (1 + increment), 2)
            else:
                # Less or equal: subtract up to 20%
                decrement = r.uniform(0, 0.2)
                coa_value = round(data[2] * (1 - decrement), 2)

            coa_cell_index = "E" + str(3 + i)
            coa_cell = self.ws[coa_cell_index]
            self.set(coa_cell, int(coa_value), "blank")

            # Calculate Capital Gain (cg): consideration_received - cost_of_acquisition
            cg_cell_index = "G" + str(3 + i)
            cg_cell = self.ws[cg_cell_index]
            cg_cell_formula = "=F" + str(3 +i) +"-E"+ str(3+i) + ""

            if (cr_value- coa_value  <= 0):
                self.set(cg_cell, cg_cell_formula, "light_red")
            else:
                self.set(cg_cell, cg_cell_formula, "green")

        """################# Total Capital Gain #################"""
        # Calculate the number of data rows
        num_rows = len(df)

        # Total Capital Gain
        tcg_formula = "=SUM(G3:G" + str(3 + num_rows -1) + ")"
        self.set(tcg_value, tcg_formula, "dark_red_h")

        # Total Cost of Acquisition
        tcoa_formula = "=SUM(E3:E" + str(3 + num_rows -1) + ")"
        self.set(tcoa_value, tcoa_formula, "light_red_h")

        # Total Consideration Received
        tcr_formula = "=SUM(F3:F" + str(3 + num_rows -1) + ")"
        self.set(tcr_value, tcr_formula, "green_h")

        self.workbook.close()
        self.workbook.save(file_name)


if __name__ == "__main__":
    # Create an instance of CSVProcessor
    test = ExcelProcessor()

    from csv_processor import CSVProcessor
    import glob
    import os

    test_folder = "test/cryptodata"
    file_list = glob.glob(os.path.join(test_folder, "*.csv"))
    print("Files found:", file_list)
    df = CSVProcessor().combine_csvs(file_list)

    test.make_dashboard("test/cryptodata/Form-16 .xlsx", df)
